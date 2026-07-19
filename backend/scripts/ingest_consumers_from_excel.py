# -*- coding: utf-8 -*-
"""从 Excel 批量录入 monitor_consumers（per-user_id 粒度）。

Excel 格式：首行表头，A 列=客户名(ai_consumer)，B 列=用户ID(customer_code/user_id)。
一行 = (ai_consumer, customer_code)，一个客户名可对应多个 user_id（占多行）。

前置：先跑 scripts/migrate_consumer_to_user_id.py 把表结构改成 customer_code UNIQUE NOT NULL。
本脚本幂等：先清空 monitor_consumers 再批量插入（含一条 __all__ 全客户汇总行，enabled=False）。

用法（backend 目录下）：
    python scripts/ingest_consumers_from_excel.py [--xlsx 路径] [config_name]
    缺省 xlsx = 仓库根目录下「客户名称uidmapping.xlsx」
"""
from __future__ import annotations

import argparse
import sys
from collections import Counter
from pathlib import Path

BACKEND_DIR = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(BACKEND_DIR))

from app.config import CONFIG_MAP, DevConfig  # noqa: E402


class IngestCfg(DevConfig):
    SCHEDULER_ENABLED = False
    USAGE_HOURLY_AGGREGATE_ENABLED = False
    RESOURCE_MONITOR_MODE = "mock"


CONFIG_MAP.setdefault("ingest_consumer", IngestCfg)

GLOBAL_AI_CONSUMER = "__all__"
DEFAULT_XLSX = BACKEND_DIR.parent / "客户名称uidmapping.xlsx"


def load_rows(xlsx: Path) -> list[tuple[str, str]]:
    from openpyxl import load_workbook
    wb = load_workbook(xlsx, data_only=True)
    ws = wb.active
    rows: list[tuple[str, str]] = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:  # 跳过表头
            continue
        name = row[0] if len(row) > 0 else None
        uid = row[1] if len(row) > 1 else None
        name = str(name).strip() if name is not None else ""
        uid = str(uid).strip() if uid is not None else ""
        if not name or not uid:
            continue
        rows.append((name, uid))
    return rows


def main() -> int:
    parser = argparse.ArgumentParser(description="从 Excel 录入 monitor_consumers（per-user_id）")
    parser.add_argument("--xlsx", type=Path, default=DEFAULT_XLSX, help="Excel 路径")
    parser.add_argument("config_name", nargs="?", default="ingest_consumer")
    args = parser.parse_args()

    if not args.xlsx.exists():
        raise SystemExit(f"Excel 不存在: {args.xlsx}")

    rows = load_rows(args.xlsx)
    print(f"[load] {args.xlsx.name}: {len(rows)} 行")

    # 校验：customer_code 全局唯一、非空
    codes = [c for _, c in rows]
    dup = {c: n for c, n in Counter(codes).items() if n > 1}
    if dup:
        raise SystemExit(f"customer_code 有重复: {dup}")
    names = [n for n, _ in rows]
    print(f"[load] distinct 客户名 {len(set(names))}，distinct user_id {len(set(codes))}")
    multi = {n: cnt for n, cnt in Counter(names).items() if cnt > 1}
    print(f"[load] 多 user_id 的客户 {len(multi)} 个："
          + "、".join(f"{n}×{c}" for n, c in list(multi.items())[:8])
          + (" ..." if len(multi) > 8 else ""))

    from app import create_app  # noqa: E402
    from app.extensions import db  # noqa: E402
    from app.models import MonitorConsumer  # noqa: E402

    app = create_app(args.config_name)
    with app.app_context():
        n_old = db.session.query(MonitorConsumer).count()
        db.session.query(MonitorConsumer).delete()  # 幂等：清空后重灌
        objs = [
            MonitorConsumer(ai_consumer=name, customer_code=uid,
                           customer_name=name, level="B", enabled=True)
            for name, uid in rows
        ]
        # 保留 __all__ 全客户汇总口径（enabled=False，不参与逐客户采集）
        objs.append(MonitorConsumer(ai_consumer=GLOBAL_AI_CONSUMER,
                                    customer_code=GLOBAL_AI_CONSUMER,
                                    customer_name="全客户汇总",
                                    level="B", enabled=False))
        db.session.bulk_save_objects(objs)
        db.session.commit()

        total = db.session.query(MonitorConsumer).count()
        enabled = db.session.query(MonitorConsumer).filter(MonitorConsumer.enabled.is_(True)).count()
        print(f"[ingest] 清空旧 {n_old} 行，录入 {total} 行（enabled={enabled}，含 1 条 __all__）")
        print(f"[ingest] distinct customer_code={db.session.query(MonitorConsumer).distinct(MonitorConsumer.customer_code).count()}")
        all_row = db.session.query(MonitorConsumer).filter(
            MonitorConsumer.customer_code == GLOBAL_AI_CONSUMER).first()
        print(f"[ingest] __all__ 行: ai_consumer={all_row.ai_consumer} enabled={all_row.enabled}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
