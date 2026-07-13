"""录入分客户小时级跑量到 customer_usage_hourly（来源：模型计量使用量明细 xlsx）。

口径：
- 明细「数据时间」已是整点小时；按 (客户, 数据时间, 模型, provider) 聚合求和 token 指标，
  落成一条 hourly（自然键与表一致）。
- 仅保留：客户名 ∈ monitor_consumers（存在的客户）且 provider ∈ provider_mappings（存在的 provider）。
- model 统一规范化为小写；model_source/data_source 取明细原值；user_id 暂空。

只影响明细涉及的这批真实客户的行（先删这批客户旧 hourly 再写），不动 __all__ 等其它行。

用法（backend 目录下）：
    python scripts/ingest_usage.py [路径/模型计量使用量明细.xlsx]
"""
import sys
from datetime import datetime, date
from pathlib import Path

import openpyxl

BACKEND_DIR = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(BACKEND_DIR))
REPO_ROOT = BACKEND_DIR.parent

from app import create_app  # noqa: E402
from app.extensions import db  # noqa: E402
from app.utils.model_name import normalize_model_name  # noqa: E402
from app.models import MonitorConsumer, ProviderMapping, CustomerUsageHourly  # noqa: E402

DETAIL_XLSX = Path(sys.argv[1]) if len(sys.argv) > 1 else REPO_ROOT / "模型计量使用量明细_20260714_004155.xlsx"


def _to_int(v):
    if v in (None, ""):
        return 0
    try:
        return int(v)
    except (TypeError, ValueError):
        return int(float(v))


def _to_dt(v):
    if isinstance(v, datetime):
        return v
    if isinstance(v, date):
        return datetime(v.year, v.month, v.day)
    return datetime.strptime(str(v).strip(), "%Y-%m-%d %H:%M:%S")


def _to_date(v):
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    return datetime.strptime(str(v).strip(), "%Y-%m-%d").date()


def load_detail(path: Path, name_to_id: dict[str, int], providers: set[str]):
    """读明细并按 (customer_id, data_time, model, provider) 聚合，返回聚合行列表。"""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    it = ws.iter_rows(values_only=True)
    header = [str(h) for h in next(it)]
    idx = {h: i for i, h in enumerate(header)}

    agg: dict[tuple, dict] = {}
    total, kept, skip_customer, skip_provider = 0, 0, 0, 0
    for r in it:
        total += 1
        name = r[idx["客户名"]]
        provider = r[idx["provider"]]
        name = name.strip() if isinstance(name, str) else name
        provider = provider.strip() if isinstance(provider, str) else provider
        cid = name_to_id.get(name)
        if cid is None:
            skip_customer += 1
            continue
        if provider not in providers:
            skip_provider += 1
            continue
        kept += 1
        dt = _to_dt(r[idx["数据时间"]])
        model = normalize_model_name(r[idx["模型"]])
        key = (cid, dt, model, provider)
        row = agg.get(key)
        if row is None:
            row = agg[key] = dict(
                customer_id=cid, customer_name=name, user_id="",
                data_time=dt, stat_date=_to_date(r[idx["日期"]]),
                model=model, provider=provider,
                model_source=r[idx["模型来源"]], data_source=r[idx["数据来源"]],
                output_token=0, cache_token=0, cache_miss_token=0,
                total_input=0, input_output=0,
                status=r[idx["状态"]], account_type=r[idx["账户类型"]],
                department=r[idx["部门"]], business_owner=r[idx["商务负责人"]],
                industry=r[idx["行业"]],
            )
        row["output_token"] += _to_int(r[idx["outputToken"]])
        row["cache_token"] += _to_int(r[idx["cacheToken"]])
        row["cache_miss_token"] += _to_int(r[idx["cacheMissToken"]])
        row["total_input"] += _to_int(r[idx["总输入"]])
        row["input_output"] += _to_int(r[idx["输入+输出"]])
    wb.close()
    print(f"[filter] 明细 {total} 行：保留 {kept}，跳过(非在册客户) {skip_customer}，"
          f"跳过(provider 不在 provider_mappings) {skip_provider}；聚合成 {len(agg)} 条小时行")
    return list(agg.values())


def main():
    print(f"明细文件: {DETAIL_XLSX}")
    app = create_app("dev")
    with app.app_context():
        db.create_all()
        name_to_id = {c.customer_name: c.id for c in db.session.query(MonitorConsumer).all()
                      if c.customer_name}
        providers = {p.provider for p in db.session.query(ProviderMapping).all() if p.provider}
        print(f"[ref] 在册客户 {len(name_to_id)} 个；provider_mappings provider {len(providers)} 个")

        rows = load_detail(DETAIL_XLSX, name_to_id, providers)
        if not rows:
            print("[done] 无匹配数据可录入")
            return

        # 幂等：先删这批客户的旧 hourly 行（不动 __all__ 等其它客户）
        cids = {r["customer_id"] for r in rows}
        deleted = CustomerUsageHourly.query.filter(
            CustomerUsageHourly.customer_id.in_(cids)).delete(synchronize_session=False)
        db.session.bulk_save_objects([CustomerUsageHourly(**r) for r in rows])
        db.session.commit()
        print(f"[usage] 删除旧行 {deleted}，写入分客户 hourly {len(rows)} 行（覆盖 {len(cids)} 个客户）")
        print(f"[done] customer_usage_hourly 现有总行数：{db.session.query(CustomerUsageHourly).count()}")


if __name__ == "__main__":
    main()
