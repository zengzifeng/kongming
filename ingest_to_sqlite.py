#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""按 realtime / time_period 两个求解器实际读取的表与字段，将
「平台输入.xlsx」+「模型计量使用量明细（分时跑量）」录入本地 SQLite。

目标库：backend/instance/kongming.db（Flask dev 配置使用的库路径）
表名/列名严格对齐 backend/app/models 下的 ORM 定义及 scripts/ingest_*.py 的转换口径：

  平台输入.自建集群信息 -> cluster_resources        (ClusterResource)
  平台输入.供应商信息   -> vendor_quotas            (VendorQuota)
  平台输入.列表价       -> model_list_prices        (ModelListPrice)
  平台输入.售卖         -> customer_sell_discounts  (CustomerSellDiscount)
  明细.Sheet2(过滤)     -> customer_usage_hourly    (CustomerUsageHourly)
  售卖客户名单          -> customers                (Customer)

因本机仅有 Python 3.9（后端需 3.11+，ORM 的 `Mapped[str|None]` 无法在 3.9 运行），
此脚本用标准库 sqlite3 复刻同一套 schema 与转换，产物与官方 ingest 脚本一致。
"""
import json
import os
import re
import sqlite3
from datetime import datetime

BASE = os.path.dirname(os.path.abspath(__file__))
PLATFORM_XLSX = os.path.join(BASE, '平台输入.xlsx')
DETAIL_XLSX = os.path.join(BASE, '模型计量使用量明细_20260709_232212.xlsx')
DB_PATH = os.path.join(BASE, 'backend', 'instance', 'kongming.db')

# 平台输入无日期字段，与实跑数据(2026-07-07)对齐的快照/生效日
SNAPSHOT_DATE = '2026-07-07'
NOW = '2026-07-07 00:00:00'   # created_at/updated_at 固定值
WTPM = 10000                  # 承接能力/供应商总量单位「万」-> 绝对量

import openpyxl


def normalize_model_name(raw):
    """对齐 app/utils/model_name.py：去首尾+折叠内部空白+小写。"""
    if raw is None:
        return ''
    return ' '.join(str(raw).split()).lower()


def _num(v):
    if v in (None, ''):
        return 0.0
    return float(v)


def _int(v):
    if v in (None, ''):
        return 0
    try:
        return int(v)
    except (TypeError, ValueError):
        return int(float(v))


def _to_dt(v):
    if isinstance(v, datetime):
        return v.strftime('%Y-%m-%d %H:%M:%S')
    return datetime.strptime(str(v).strip(), '%Y-%m-%d %H:%M:%S').strftime('%Y-%m-%d %H:%M:%S')


def _to_date(v):
    if isinstance(v, datetime):
        return v.strftime('%Y-%m-%d')
    return datetime.strptime(str(v).strip(), '%Y-%m-%d').strftime('%Y-%m-%d')


def read_sheet(wb, name):
    ws = wb[name]
    rows = list(ws.iter_rows(values_only=True))
    header = [str(h).strip() if h is not None else None for h in rows[0]]
    ncol = max((i for i, h in enumerate(header) if h), default=-1) + 1
    out = []
    for r in rows[1:]:
        if all(v is None for v in r[:ncol]):
            continue
        out.append({header[i]: r[i] for i in range(ncol)})
    return out


# ---------------------------------------------------------------- schema
DDL = """
CREATE TABLE customers (
  id INTEGER PRIMARY KEY AUTOINCREMENT,
  created_at TEXT NOT NULL, updated_at TEXT NOT NULL,
  customer_code TEXT UNIQUE NOT NULL, name TEXT NOT NULL,
  level TEXT NOT NULL DEFAULT 'B', strategic_tag TEXT,
  paid_amount_total NUMERIC DEFAULT 0, signed_at TEXT, extra_json TEXT
);
CREATE TABLE customer_usage_hourly (
  id INTEGER PRIMARY KEY AUTOINCREMENT,
  created_at TEXT NOT NULL, updated_at TEXT NOT NULL,
  customer_id INTEGER NOT NULL, customer_name TEXT NOT NULL,
  user_id TEXT NOT NULL, key_id TEXT NOT NULL,
  data_time TEXT NOT NULL, stat_date TEXT NOT NULL,
  phase TEXT NOT NULL, model TEXT NOT NULL, provider TEXT NOT NULL,
  model_source TEXT, data_source TEXT,
  output_token INTEGER DEFAULT 0, cache_token INTEGER DEFAULT 0,
  cache_miss_token INTEGER DEFAULT 0, total_input INTEGER DEFAULT 0,
  input_output INTEGER DEFAULT 0,
  creation_cache_1h_token INTEGER DEFAULT 0, creation_cache_5m_token INTEGER DEFAULT 0,
  web_search_fc_count INTEGER DEFAULT 0, av_duration NUMERIC DEFAULT 0,
  status TEXT, account_type TEXT, department TEXT, business_owner TEXT, industry TEXT,
  FOREIGN KEY(customer_id) REFERENCES customers(id),
  CONSTRAINT uq_usage_hourly_natural_key UNIQUE
    (customer_id,user_id,key_id,data_time,stat_date,model,provider,phase)
);
CREATE INDEX ix_usage_hourly_customer_time ON customer_usage_hourly(customer_id,data_time);
CREATE INDEX ix_usage_hourly_stat_date ON customer_usage_hourly(stat_date);
CREATE TABLE customer_sell_discounts (
  id INTEGER PRIMARY KEY AUTOINCREMENT,
  created_at TEXT NOT NULL, updated_at TEXT NOT NULL,
  customer_id INTEGER NOT NULL, customer_name TEXT NOT NULL,
  model_name TEXT NOT NULL, sell_discount NUMERIC DEFAULT 0,
  effective_from TEXT NOT NULL, effective_to TEXT,
  FOREIGN KEY(customer_id) REFERENCES customers(id),
  CONSTRAINT uq_sell_discount_customer_model_effective UNIQUE
    (customer_id,model_name,effective_from)
);
CREATE INDEX ix_sell_discount_model ON customer_sell_discounts(model_name);
CREATE TABLE model_list_prices (
  id INTEGER PRIMARY KEY AUTOINCREMENT,
  created_at TEXT NOT NULL, updated_at TEXT NOT NULL,
  model_name TEXT NOT NULL,
  input_cache_hit_price NUMERIC DEFAULT 0, input_cache_miss_price NUMERIC DEFAULT 0,
  output_price NUMERIC DEFAULT 0,
  effective_from TEXT NOT NULL, effective_to TEXT,
  CONSTRAINT uq_model_price_effective UNIQUE (model_name,effective_from)
);
CREATE INDEX ix_model_list_prices_model_name ON model_list_prices(model_name);
CREATE TABLE cluster_resources (
  id INTEGER PRIMARY KEY AUTOINCREMENT,
  created_at TEXT NOT NULL, updated_at TEXT NOT NULL,
  snapshot_date TEXT NOT NULL,
  cluster_name TEXT NOT NULL, deployed_model TEXT NOT NULL, primary_customer TEXT,
  machine_count INTEGER DEFAULT 0, tpm_per_machine NUMERIC DEFAULT 0,
  total_capacity_tpm NUMERIC DEFAULT 0,
  peak_tpm_d1_23_24 NUMERIC DEFAULT 0, peak_tpm_d2_23_24 NUMERIC DEFAULT 0,
  peak_tpm_d3_23_24 NUMERIC DEFAULT 0, peak_tpm_idle NUMERIC DEFAULT 0,
  idle_redundant_tpm NUMERIC DEFAULT 0, idle_redundant_machines INTEGER DEFAULT 0,
  peak_tpm_busy NUMERIC DEFAULT 0, busy_redundant_tpm NUMERIC DEFAULT 0,
  busy_redundant_machines INTEGER DEFAULT 0,
  current_tpm NUMERIC DEFAULT 0, current_redundant_tpm NUMERIC DEFAULT 0,
  current_redundant_machines INTEGER DEFAULT 0, raw_json TEXT
);
CREATE INDEX ix_cluster_resources_snapshot_date ON cluster_resources(snapshot_date);
CREATE INDEX ix_cluster_resources_cluster_name ON cluster_resources(cluster_name);
CREATE INDEX ix_cluster_resources_deployed_model ON cluster_resources(deployed_model);
CREATE TABLE vendor_quotas (
  id INTEGER PRIMARY KEY AUTOINCREMENT,
  created_at TEXT NOT NULL, updated_at TEXT NOT NULL,
  vendor TEXT NOT NULL, model TEXT NOT NULL,
  quota_tpm NUMERIC DEFAULT 0, actual_tpm NUMERIC DEFAULT 0,
  actual_redundant_tpm NUMERIC DEFAULT 0,
  unit_cost NUMERIC DEFAULT 0, unit_price NUMERIC DEFAULT 0,
  purchase_discount NUMERIC DEFAULT 0,
  effective_from TEXT NOT NULL, effective_to TEXT,
  status TEXT NOT NULL DEFAULT 'active', contact TEXT, notes TEXT, raw_json TEXT,
  CONSTRAINT uq_vendor_model_effective UNIQUE (vendor,model,effective_from)
);
CREATE INDEX ix_vendor_quotas_vendor ON vendor_quotas(vendor);
CREATE INDEX ix_vendor_quotas_model ON vendor_quotas(model);
CREATE INDEX ix_vendor_quotas_status ON vendor_quotas(status);
"""


def load_valid_customers(wb):
    """售卖 sheet 的客户名单（realtime/period 需求 universe 之外的客户被过滤）。"""
    names = set()
    for r in read_sheet(wb, '售卖'):
        n = r.get('客户名称')
        if n and str(n).strip():
            names.add(str(n).strip())
    return names


def main():
    if os.path.exists(DB_PATH):
        os.remove(DB_PATH)
    conn = sqlite3.connect(DB_PATH)
    conn.executescript(DDL)
    cur = conn.cursor()

    pf = openpyxl.load_workbook(PLATFORM_XLSX, data_only=True)
    valid = load_valid_customers(pf)

    # ---- customers：按售卖名单，sorted 顺序分配 C0001..（对齐 ingest_usage）----
    name_to_id = {}
    for i, name in enumerate(sorted(valid), start=1):
        code = f'C{i:04d}'
        cur.execute(
            "INSERT INTO customers(created_at,updated_at,customer_code,name,level,"
            "paid_amount_total,extra_json) VALUES(?,?,?,?,?,?,?)",
            (NOW, NOW, code, name, 'B', 0, '{}'))
        name_to_id[name] = cur.lastrowid
    print(f'[customers]              {len(name_to_id)} 行')

    # ---- cluster_resources（自建集群信息）----
    n = 0
    for r in read_sheet(pf, '自建集群信息'):
        mc = _int(r['部署机器台数'])
        tpm_per = _num(r['单台承接能力TPM']) * WTPM
        # 总承接能力为公式(=台数×单台)，data_only 可能为空 -> 用台数×单台兜底
        total = _num(r['总承接能力TPM']) * WTPM or mc * tpm_per
        cur.execute(
            "INSERT INTO cluster_resources(created_at,updated_at,snapshot_date,cluster_name,"
            "deployed_model,machine_count,tpm_per_machine,total_capacity_tpm,raw_json) "
            "VALUES(?,?,?,?,?,?,?,?,?)",
            (NOW, NOW, SNAPSHOT_DATE, str(r['自建集群名称']),
             normalize_model_name(r['部署模型名称']), mc, tpm_per, total,
             json.dumps({'provider': r['provider'],
                         '单台承接能力_wTPM': r['单台承接能力TPM'],
                         '总承接能力_wTPM': r['总承接能力TPM'],
                         'source': '平台输入.自建集群信息'}, ensure_ascii=False)))
        n += 1
    print(f'[cluster_resources]      {n} 行')

    # ---- vendor_quotas（供应商信息）：vendor=供应商名称, model 规范化 ----
    n = 0
    for r in read_sheet(pf, '供应商信息'):
        cur.execute(
            "INSERT INTO vendor_quotas(created_at,updated_at,vendor,model,quota_tpm,"
            "purchase_discount,effective_from,status,raw_json) VALUES(?,?,?,?,?,?,?,?,?)",
            (NOW, NOW, str(r['供应商名称']), normalize_model_name(r['模型名称']),
             _num(r['供应商总量(W)']) * WTPM, _num(r['采购折扣']), SNAPSHOT_DATE, 'active',
             json.dumps({'供应商总量_W': r['供应商总量(W)'], 'source': '平台输入.供应商信息'},
                        ensure_ascii=False)))
        n += 1
    print(f'[vendor_quotas]          {n} 行')

    # ---- model_list_prices（列表价）----
    n = 0
    for r in read_sheet(pf, '列表价'):
        if not r.get('模型名称'):
            continue
        cur.execute(
            "INSERT INTO model_list_prices(created_at,updated_at,model_name,"
            "input_cache_hit_price,input_cache_miss_price,output_price,effective_from) "
            "VALUES(?,?,?,?,?,?,?)",
            (NOW, NOW, normalize_model_name(r['模型名称']), _num(r['输入命中列表价']),
             _num(r['输入未命中列表价']), _num(r['输出列表价']), SNAPSHOT_DATE))
        n += 1
    print(f'[model_list_prices]      {n} 行')

    # ---- customer_sell_discounts（售卖）：按客户名连接 customers ----
    n, missing = 0, []
    for r in read_sheet(pf, '售卖'):
        name = str(r['客户名称']).strip()
        cid = name_to_id.get(name)
        if cid is None:
            missing.append(name)
            continue
        cur.execute(
            "INSERT INTO customer_sell_discounts(created_at,updated_at,customer_id,"
            "customer_name,model_name,sell_discount,effective_from) VALUES(?,?,?,?,?,?,?)",
            (NOW, NOW, cid, name, normalize_model_name(r['模型名称']),
             _num(r['售卖折扣']), SNAPSHOT_DATE))
        n += 1
    print(f'[customer_sell_discounts] {n} 行' + (f'（跳过未匹配 {missing}）' if missing else ''))
    pf.close()

    # ---- customer_usage_hourly（明细，过滤非售卖客户）----
    wb = openpyxl.load_workbook(DETAIL_XLSX, read_only=True, data_only=True)
    ws = wb['Sheet2']
    it = ws.iter_rows(values_only=True)
    header = [str(h) for h in next(it)]
    idx = {h: i for i, h in enumerate(header)}
    kept = dropped = 0
    batch = []
    cols = ("customer_id,customer_name,user_id,key_id,data_time,stat_date,phase,model,"
            "provider,model_source,data_source,output_token,cache_token,cache_miss_token,"
            "total_input,input_output,creation_cache_1h_token,creation_cache_5m_token,"
            "web_search_fc_count,av_duration,status,account_type,department,business_owner,industry")
    sql = (f"INSERT OR IGNORE INTO customer_usage_hourly(created_at,updated_at,{cols}) "
           f"VALUES(?,?,{','.join('?'*25)})")
    for r in it:
        name = r[idx['客户名']]
        name = str(name).strip() if name is not None else ''
        if name not in valid:
            dropped += 1
            continue
        batch.append((
            NOW, NOW, name_to_id[name], name,
            str(r[idx['用户ID']]), str(r[idx['key_id']]),
            _to_dt(r[idx['数据时间']]), _to_date(r[idx['日期']]),
            str(r[idx['阶段']]), normalize_model_name(r[idx['模型']]), str(r[idx['provider']]),
            r[idx['模型来源']], r[idx['数据来源']],
            _int(r[idx['outputToken']]), _int(r[idx['cacheToken']]), _int(r[idx['cacheMissToken']]),
            _int(r[idx['总输入']]), _int(r[idx['输入+输出']]),
            _int(r[idx['creationCache1hToken']]), _int(r[idx['creationCache5mToken']]),
            _int(r[idx['webSearchFcCount']]), _num(r[idx['音视频时长']]),
            r[idx['状态']], r[idx['账户类型']], r[idx['部门']], r[idx['商务负责人']], r[idx['行业']]))
        kept += 1
        if len(batch) >= 1000:
            cur.executemany(sql, batch); batch.clear()
    if batch:
        cur.executemany(sql, batch)
    wb.close()
    print(f'[customer_usage_hourly]  保留 {kept} 行，过滤 {dropped} 行')

    conn.commit()
    print('\n=== 各表行数 ===')
    for (t,) in cur.execute("SELECT name FROM sqlite_master WHERE type='table' ORDER BY name"):
        c = cur.execute(f'SELECT COUNT(*) FROM "{t}"').fetchone()[0]
        print(f'  {t:24s} {c}')
    conn.close()
    print(f'\n数据库: {DB_PATH}')


if __name__ == '__main__':
    main()
